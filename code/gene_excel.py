
from main import *
from predict import *
from split_wrap import *
from scipy.linalg import expm
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side

Inf = 1e10
def gene_diff_record(cancer_name):
    dir_path = '../tem_data/{}_records/'.format(cancer_name)
    if not Path(dir_path).is_dir():
        Path(dir_path).mkdir()
    _, y = construct_and_return_loc_net()
    for i in tqdm(range(11, 12)):
        i /= 10
        nor_proba = np.load('../tem_data/{}_probas/probas_nor_tau_{}.npy'.format(cancer_name, i))
        ill_proba = np.load('../tem_data/{}_probas/probas_ill_tau_{}.npy'.format(cancer_name, i))
        nor_proba = proba_scaling(nor_proba)
        ill_proba = proba_scaling(ill_proba)

        with open('../data/uni_ids_with_loc.txt') as f:
            proteins = f.read().split()
        with open('../data/location_name.txt') as f:
            locations = f.read().split('\n')#位置列表，12个
            locations = np.array(locations)
        with open(dir_path + '{}_27{}.csv'.format(cancer_name, i), 'w') as f:
            f.write(' , , ,')
            f.write(','.join(locations.tolist()))
            f.write(','.join(locations.tolist()))
            f.write(','.join(locations.tolist()))
            f.write('\n')
            id2gene = gene_mapping('id')
            id2entry = gene_mapping('id2entry')
            for row in range(len(y)):
                protein_id = proteins[row]
                entry = id2entry[protein_id]
                gene_names = id2gene[protein_id]
                nor_p = [str(round(x, 3)) for x in nor_proba[row]]
                ill_p = [str(round(x, 3)) for x in ill_proba[row]]
                diff = (ill_proba[row] - nor_proba[row]) / nor_proba[row]
                diff[np.isnan(diff)] = 0
                str_diff = ['{:.2%}'.format(x) for x in diff]
                f.write(
                        ','.join([
                            protein_id, 
                            entry, 
                            ' '.join(gene_names).ljust(30),
                            *nor_p,
                            *ill_p,
                            *str_diff,
                            ])
                        )
                f.write('\n')

def gene_diff_record():
    _, y = construct_and_return_loc_net()
    tau = 1.1
    with open('../data/uni_ids_with_loc.txt') as f:
        proteins = f.read().split()
    with open('../data/location_name.txt') as f:
        locations = f.read().strip().split('\n')#位置列表，12个
        locations = np.array(locations)
    wb = Workbook(write_only=True)
    for cancer_name in cancer_names:
        nor_proba = np.load('../tem_data/{}_probas/probas_nor_tau_{}.npy'.format(cancer_name, tau))
        ill_proba = np.load('../tem_data/{}_probas/probas_ill_tau_{}.npy'.format(cancer_name, tau))
        nor_proba = proba_scaling(nor_proba)
        ill_proba = proba_scaling(ill_proba)
        diff = (ill_proba - nor_proba) / nor_proba
        diff[np.isnan(diff)] = 0
        diff[np.isinf(diff)] = Inf# 替换inf
        ws = wb.create_sheet('{}_all'.format(cancer_name))
        ws.append([])
        ws.append([])
        ws.append([None, None, None, None] + locations.tolist() * 3)
        id2gene = gene_mapping('id')
        id2entry = gene_mapping('id2entry')
        for row in range(len(y)):
            protein_id = proteins[row]
            entry = id2entry[protein_id]
            gene_names = ' '.join(id2gene[protein_id])

            nor_p = nor_proba[row]
            ill_p = ill_proba[row]
            diff_p = diff[row]
            '''
            if np.isinf(diff_p).any():
                diff_p = ['Inf' if np.isinf(x) else x for x in diff_p]
                '''
            ws.append([None, protein_id, entry, gene_names, *nor_p, *ill_p, *diff_p])
    wb.save('../tem_data/all.xlsx')

def gene_xlsx():
    with open('../data/uni_ids_with_loc.txt') as f:
        proteins = f.read().split()
    with open('../data/location_name.txt') as f:
        locations = f.read().strip().split('\n')#位置列表，12个
        locations = np.array(locations)
    _, y = construct_and_return_loc_net()
    id2gene = gene_mapping('id')
    id2entry = gene_mapping('id2entry')

    #wb = Workbook(write_only=True)
    wb = Workbook()
    #格式
    ft = Font(name='Times New Roman')
    bd = Side(border_style='thin')

    for cancer_name in cancer_names:
        dir_path = '../tem_data/{}_records/'.format(cancer_name)
        if not Path(dir_path).is_dir():
            Path(dir_path).mkdir()
        tau = 1.1
        nor_proba = np.load('../tem_data/{}_probas/probas_nor_tau_{}.npy'.format(cancer_name, tau))
        ill_proba = np.load('../tem_data/{}_probas/probas_ill_tau_{}.npy'.format(cancer_name, tau))
        nor_proba = proba_scaling(nor_proba)
        ill_proba = proba_scaling(ill_proba)
        relative_diff = (ill_proba - nor_proba) / nor_proba
        relative_diff[np.isnan(relative_diff)] = 0
        relative_diff[np.isinf(relative_diff)] = Inf

        
        for location_index in range(len(locations)):
            location = locations[location_index]
            ws = wb.create_sheet('{}_{}'.format(cancer_name, location))
            ws.append([])
            ws.append([None, 'Uniprot AC', 'Uniprot ID', 'Gene names', 'Relative change ({})'.format(location)])
            for row in range(len(y)):
                protein_id = proteins[row]
                entry = id2entry[protein_id]
                gene_names = id2gene[protein_id]
                diff = relative_diff[row, location_index]
                str_diff = '{:.2%}'.format(diff)
                ws.append([None, protein_id, entry, ' '.join(gene_names).ljust(30), diff])

            for col in 'BCDE':
                #ws['{}2'.format(col)].style = style_header
                cell = ws['{}2'.format(col)]
                cell.font = ft
                cell.border = Border(top=bd, bottom=bd)
                cell = ws['{}6463'.format(col)]
                cell.font = ft
                cell.border = Border(bottom=bd)
            for col in 'BCD':
                for row in range(3, 6463):
                    cell = ws['{}{}'.format(col, row)]
                    cell.font = ft
            for row in range(3, 6464):
                cell = ws['E{}'.format(row)]
                cell.font = ft
                #cell.number_format = 'Percent'

    wb.save('../tem_data/39.xlsx')

def max_min_xlsx():
    with open('../data/uni_ids_with_loc.txt') as f:
        proteins = f.read().split()
    with open('../data/location_name.txt') as f:
        locations = f.read().strip().split('\n')#位置列表，12个
        locations = np.array(locations)
    _, y = construct_and_return_loc_net()
    id2gene = gene_mapping('id')
    id2entry = gene_mapping('id2entry')

    #wb = Workbook(write_only=True)
    wb = Workbook()
    #格式
    ft = Font(name='Times New Roman')
    bd = Side(border_style='thin')

    for cancer_name in cancer_names:
        dir_path = '../tem_data/{}_records/'.format(cancer_name)
        if not Path(dir_path).is_dir():
            Path(dir_path).mkdir()
        tau = 1.1
        nor_proba = np.load('../tem_data/{}_probas/probas_nor_tau_{}.npy'.format(cancer_name, tau))
        ill_proba = np.load('../tem_data/{}_probas/probas_ill_tau_{}.npy'.format(cancer_name, tau))
        nor_proba = proba_scaling(nor_proba)
        ill_proba = proba_scaling(ill_proba)
        relative_diff = (ill_proba - nor_proba) / nor_proba
        relative_diff[np.isnan(relative_diff)] = 0
        relative_diff[np.isinf(relative_diff)] = Inf
        max_diff = relative_diff.max(1)
        max_location_index = relative_diff.argmax(1)
        min_diff = relative_diff.min(1)
        min_location_index = relative_diff.argmin(1)
        
        for arr_sign in ('max', 'min'):
            if arr_sign == 'max':
                a = max_diff
                indices = max_location_index
            elif arr_sign == 'min':
                a = min_diff
                indices = min_location_index
            ws = wb.create_sheet('{}_{}'.format(cancer_name, arr_sign))
            ws.append([])
            ws.append([None, 'Uniprot AC', 'Uniprot ID', 'Gene names', 'Max relative change', 'Location'])
            for row in range(len(y)):
                protein_id = proteins[row]
                entry = id2entry[protein_id]
                gene_names = id2gene[protein_id]
                diff = a[row]
                locatin = locations[indices[row]]
                ws.append([None, protein_id, entry, ' '.join(gene_names).ljust(30), diff, locatin])

    wb.save('../tem_data/max_min.xlsx')
if __name__ == '__main__':
    cancer_names = [
            'hepatitis',
            'leukemia',
            'breast'
            ]
    gene_diff_record()
    gene_xlsx()
    max_min_xlsx()
